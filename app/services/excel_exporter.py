from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.models.test_case import TestCase


def build_excel(cases: list[TestCase]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Test Cases"

    headers = ["ID", "Title", "Precondition", "Steps", "Expected", "Type"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    for case in cases:
        sheet.append(
            [
                case.id,
                case.title,
                case.precondition,
                "\n".join(case.steps),
                "\n".join(case.expected),
                case.type.value,
            ]
        )

    widths = [14, 36, 42, 54, 54, 18]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream

